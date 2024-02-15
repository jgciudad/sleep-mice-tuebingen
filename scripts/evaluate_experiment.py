#!/usr/bin/env python

import argparse
import sys
from importlib import import_module
from os.path import basename, join, dirname, realpath, isfile
import sklearn.metrics
import numpy as np
import openpyxl

import torch
import torch.utils.data as t_data

sys.path.insert(0, realpath(join(dirname(__file__), '..')))

from base.config_loader import ConfigLoader
from base.data.dataloader import TuebingenDataloader
from base.logger import Logger
from base.evaluation.evaluate_model import evaluate
from base.evaluation.result_logger import ResultLogger


def parse():
    parser = argparse.ArgumentParser(description='evaluate exp')
    parser.add_argument('--experiment', '-e', required=True,
                        help='name of experiment to run')
    parser.add_argument('--dataset', '-d', default='valid',
                        help='dataset to evaluate model on')

    return parser.parse_args()


def evaluation(dataset, c, r, excel_path):
    """evaluates best model in experiment on given dataset"""
    logger.fancy_log('start evaluation')
    result_logger = ResultLogger(config)

    # create dataloader for given dataset, the data should not be altered in any way
    map_loader = TuebingenDataloader(config, dataset, balanced=False, augment_data=False)
    dataloader = t_data.DataLoader(map_loader, batch_size=config.BATCH_SIZE_EVAL, shuffle=False, num_workers=4)

    # create empty model from model name in config and set it's state from best model in EXPERIMENT_DIR
    model = import_module('.' + config.MODEL_NAME, 'base.models').Model(config).to(config.DEVICE).eval()
    model_file = join(config.EXPERIMENT_DIR, config.MODEL_NAME + '-best.pth')
    if isfile(model_file):
        model.load_state_dict(torch.load(model_file, map_location=torch.device('cpu'))['state_dict'])
    else:
        raise ValueError('model_file {} does not exist'.format(model_file))
    logger.logger.info('loaded model:\n' + str(model))

    # evaluate model
    labels, _ = evaluate(config, model, dataloader)

    true = labels['actual'][labels['actual'] != 3]
    predicted = labels['predicted'][labels['actual'] != 3]

    cm = sklearn.metrics.confusion_matrix(y_true=true, y_pred=predicted, labels=[1, 2, 0])
    print(cm)
    kappa = sklearn.metrics.cohen_kappa_score(y1=true, y2=predicted, labels=np.arange(len(config.STAGES[:-1])))
    print(kappa)
    
    wb = openpyxl.load_workbook(excel_path)   
    sheet = wb["Sheet1"]
    for i in range(cm.shape[0]):
        for j in range(cm.shape[1]):
            sheet.cell(row = r + i, column = c + j).value = cm[i, j]
    sheet.cell(row = r-1, column = c-3).value = kappa
    wb.save(excel_path)

    # # log/plot results
    # result_logger.log_sleep_stage_f1_scores(labels['actual'], labels['predicted'], dataset)
    # logger.logger.info('')
    # result_logger.log_confusion_matrix(labels['actual'], labels['predicted'], dataset, wo_plot=False)
    # result_logger.log_transformation_matrix(labels['actual'], labels['predicted'], dataset,
    #                                         wo_plot=False)

    # logger.fancy_log('finished evaluation')


if __name__ == '__main__':
    # args = parse()
    # config = ConfigLoader(args.experiment)  # load config from experiment

    excel_path = '/home/s202283/code/sleep-mice-tuebingen/grieger_reduced.xlsx'
    exp = 'kornum_config_it3'
    test_datasets = ['test_reduced', 'spA_scorer1', 'spA_scorer2', 'spD_scorer1', 'spD_scorer2']

    for d in test_datasets:

        if d=='test_reduced':
            c = 8
        elif d=='spA_scorer1':
            c = 19
        elif d=='spA_scorer2':
            c = 30
        elif d=='spD_scorer1':
            c = 41
        elif d=='spD_scorer2':
            c = 52

        if exp == 'kornum_config_it1':
            r = 5
        elif exp == 'kornum_config_it2':
            r = 12
        elif exp == 'kornum_config_it3':
            r = 19
        # else:
        #     r=33

        config = ConfigLoader(experiment=exp)

        logger = Logger(config)  # create wrapper for logger
        # logger.init_log_file(args, basename(__file__))  # create log file and log config, etc

        # dataset = args.dataset
        # dataset = 'test'

        logger.fancy_log('evaluate best model of experiment {} on dataset {}'.format(exp, d))
        # perform evaluation

        evaluation(d, c, r, excel_path)
